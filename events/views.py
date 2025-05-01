
from django.shortcuts import render,get_object_or_404,redirect
from django.http import JsonResponse,HttpResponse,HttpResponseRedirect

from .forms import ParticipantForm
from .models import Event, Participant,Division,Match
import openpyxl
from django.contrib.auth.decorators import login_required
from django.urls import reverse
from django.template.loader import render_to_string
from django.db.models import F
import json

#TODO

# list of registered people 
# More info to add to event




def eventDetail(request,event_id):
    event = Event.objects.filter(id=event_id)
    participants = Participant.objects.filter(event=event_id, deleted=False).order_by("weight", "first_name")
    divisions = Division.objects.filter(event=event_id).order_by("name")
    if request.method == "POST":
        form = ParticipantForm(request.POST)
        if form.is_valid():
            form.save()  
            return render(request, "confirmRegistration.html", {'success': True, 'event':event,'divisions': divisions})
        else:
            print(form.errors)
            return render(request, "eventDetail.html", {'form': form, 'errors': form.errors,'event':event,'divisions': divisions})
    
    return render(request, "eventDetail.html", {'event':event, 'participants':participants,'divisions': divisions})


def register_participant(request):
    if request.method == "POST":
        form = ParticipantForm(request.POST)
        if form.is_valid():
            form.save()  
            return render(request, "eventDetail.html", {'success': True})
        else:
            return render(request, "eventDetail.html", {'form': form, 'errors': form.errors})
    return render(request, "eventDetail.html")


@login_required
def exportEventParticipants(request,event_id):
    if request.method == 'POST':
        if not event_id:
            return JsonResponse({'error': 'Event ID is required'}, status=400)

        participants = Participant.objects.filter(event=event_id, deleted=False).order_by("weight", "first_name")

        # Prepare the Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Wolfhouse'

        # Write headers
        headers = ['Full Name',  'Weight', 'DOB', 'Years Of Training','Gym','Gender']
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)

        row_num = 2

        for participant in participants:
            birthdate_value = participant.date_of_birth.strftime('%Y-%m-%d') if participant.date_of_birth else "N/A"
            ws.cell(row=row_num, column=1, value=f'{participant.first_name} {participant.last_name}')
            ws.cell(row=row_num, column=2, value=participant.weight)
            ws.cell(row=row_num, column=3, value=birthdate_value)
            ws.cell(row=row_num, column=4, value=participant.years_of_training)
            ws.cell(row=row_num, column=5, value=participant.school_name)
            ws.cell(row=row_num, column=6, value=participant.gender)
            row_num += 1

        # Create a response object to serve the Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Event_Participants.xlsx"'

        # Save the workbook to the response
        wb.save(response)

        return response

    return JsonResponse({'error': 'Invalid request method'}, status=400)

@login_required
def delete_participant(request, percipient):
    p = get_object_or_404(Participant, id=percipient)
    if request.method == "POST":
        event_id = request.POST.get('event_id')

        
        p.deleted = True
        p.save()

        return redirect(reverse('eventDetail', kwargs={'event_id': event_id}))
    return redirect('home')

@login_required
def paid_participant(request, percipient):
    p = get_object_or_404(Participant, id=percipient)
    if request.method == "POST":
        event_id = request.POST.get('event_id')
        print(request.POST)

        
        p.payment_id = 1
        p.save()

        return redirect(reverse('eventDetail', kwargs={'event_id': event_id}))
    return redirect('home')


def match_list(request, event_id):
    # Get the event based on event_id
    event = get_object_or_404(Event, id=event_id)

    # Get all divisions related to the event
    divisions = Division.objects.filter(event=event)

    # Prepare the data structure with divisions and their matches
    divisions_with_matches = []

    for division in divisions:
        # Get all matches for the current division
        matches = Match.objects.filter(division=division)
        
        participants = Participant.objects.filter(division=division,deleted=False).annotate(
            total_points=F('points')).order_by('-total_points')

        # Add the division and its matches to the list
        divisions_with_matches.append({
            'division': division.name,  # or whatever field you want to display from Division
            'matches': [
                {
                    'participant1': {
                        'id': match.participant1.id if match.participant1 else None,
                        'name': f"{match.participant1.first_name} {match.participant1.last_name}" if match.participant1 else "Bye"
                    },
                    'participant2': {
                        'id': match.participant2.id if match.participant2 else None,
                        'name': f"{match.participant2.first_name} {match.participant2.last_name}" if match.participant2 else "Bye"
                    },
                    'winner': {
                        'id': match.winner.id,
                        'name': f"{match.winner.first_name} {match.winner.last_name}"
                    } if match.winner else None,
                    'match_id': match.id,
                    'round_number': match.round_number
                }
                for match in matches
        ],

            'participants': [
                {
                    'name': f"{p.first_name} {p.last_name}",
                    'points': p.points  # or whatever your score field is
                } for p in participants
            ]
        })
    
    # Render the partial template with the new data structure
    html = render_to_string(
        'match_list.html',
        {
            'divisions_with_matches': divisions_with_matches,
            'event': event
        },
        request=request  # ✅ This includes context processors like csrf_token
    )

    # Return the HTML as a JSON response
    return JsonResponse({'html': html})


@login_required
def update_match_winner(request):
    if request.method == "POST":
        match_id = request.POST.get('match_id')
        winner_id = request.POST.get('winner_id')
        print(match_id,winner_id)
        try:
            match = Match.objects.get(id=match_id)
            if winner_id:
                winner = Participant.objects.get(id=winner_id)
                match.winner = winner
            else:
                match.winner = None
            match.save()
            
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    else:
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    


@login_required
def update_participant_division(request):
    if request.method == "POST":
        participant_id = request.POST.get('participant_id')
        division_id = request.POST.get('division')
        event_id = request.POST.get('event_id')
        try:
            
            participant = get_object_or_404(Participant, id=participant_id)
            if division_id:
                division = get_object_or_404(Division, id=division_id)
                participant.division = division
            else:
                participant.division = None  
            
            participant.save()

            return redirect('eventDetail', event_id=event_id)

        except Exception as e:
            print(f"Error: {e}")
            
            return redirect('eventDetail', event_id=event_id)

    
    return redirect('eventDetail', event_id=event_id)
